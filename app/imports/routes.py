from __future__ import annotations

import csv
import io
import json
import re
import shutil
import time
import unicodedata
from datetime import date, datetime, timedelta, time as dt_time
from decimal import Decimal, InvalidOperation
from pathlib import Path

from flask import Blueprint, abort, current_app, flash, redirect, render_template, request, send_from_directory, url_for
from flask_login import current_user, login_required
from openpyxl import load_workbook
from sqlalchemy import and_
from werkzeug.utils import secure_filename

from app.extensions import db
from app.models import BaseOperacional, CasoDNR, ImportacaoLote

bp = Blueprint("imports", __name__, url_prefix="/importacoes")

# Os aliases abaixo contemplam o leiaute oficial enviado pela operação e
# variações comuns usadas por outras bases.
ALIASES = {
    "tbr": ["tbr", "tracking id", "tracking", "codigo rastreio", "código rastreio"],
    "cliente": ["cliente", "customer", "nome cliente", "customer name"],
    "endereco": ["endereco", "endereço", "address", "customer address"],
    "cep": ["cep mapa", "cep", "postal", "postal code", "zipcode", "zip"],
    "motorista": [
        "nome do agente de entrega", "motorista", "nome do motorista", "nome motorista",
        "agente de entrega", "driver", "driver name", "delivery associate",
    ],
    "login_utilizado": [
        "login utilizado", "login em uso", "login usado", "login", "driver login",
        "login motorista", "login do motorista", "delivery associate login", "da login",
        "login da", "usuario motorista", "usuário motorista", "login utilizado na rota",
    ],
    "login_proprio": ["login proprio", "login próprio", "proprio", "próprio"],
    "proprietario_login": ["proprietario login", "proprietário login", "dono do login", "login owner"],
    "produto": ["produto", "product", "item"],
    "categoria": ["categoria", "category", "categoria produto"],
    "valor": ["valor", "price", "preco", "preço", "valor produto"],
    "data_abertura_dnr": [
        "data de abertura do dnr", "data abertura dnr", "data de abertura",
        "abertura do dnr", "dnr open date", "open date", "data abertura",
    ],
    "data_dnr": [
        "data da entrega", "data dnr", "delivery date", "data entrega",
        "data ocorrencia", "data ocorrência", "data",
    ],
    "data_hora_entrega": [
        "data de entrega", "data/hora", "data hora", "datetime", "timestamp",
        "delivery datetime", "data e hora da entrega",
    ],
    "hora_dnr": [
        "hora da entrega", "hora", "horario", "horário", "hora dnr", "horario dnr",
        "horário dnr", "delivery time", "hora entrega", "horario entrega",
        "horário entrega", "hora ocorrencia", "hora ocorrência", "time",
    ],
    "semana": ["semana", "week", "semana dnr", "número da semana", "numero da semana"],
    "pedido": ["pedido", "order", "order id", "id pedido"],
}


def norm(value: object) -> str:
    text = str(value or "").strip().lower()
    text = unicodedata.normalize("NFKD", text)
    text = "".join(ch for ch in text if not unicodedata.combining(ch))
    text = re.sub(r"[^a-z0-9]+", " ", text)
    return re.sub(r"\s+", " ", text).strip()


def detect_columns(headers: list[str]) -> dict[str, str]:
    """Detecta colunas sem confundir 'Data de entrega' com 'data da entrega'."""
    result: dict[str, str] = {}
    remaining = list(headers)

    # O layout oficial diferencia esses campos apenas pela capitalização.
    exact_priority = {
        "data_abertura_dnr": ["Data de abertura do DNR", "data de abertura do dnr"],
        "data_hora_entrega": ["Data de entrega"],
        "data_dnr": ["data da entrega"],
        "hora_dnr": ["hora da entrega"],
        "motorista": ["Nome do agente de entrega"],
        "cep": ["CEP mapa"],
    }
    for field, names in exact_priority.items():
        for name in names:
            if name in remaining:
                result[field] = name
                remaining.remove(name)
                break

    normalized: dict[str, list[str]] = {}
    for header in remaining:
        normalized.setdefault(norm(header), []).append(header)

    for field, aliases in ALIASES.items():
        if field in result:
            continue
        for alias in aliases:
            candidates = normalized.get(norm(alias), [])
            if candidates:
                result[field] = candidates[0]
                break
    return result


def parse_decimal(value: object) -> Decimal:
    raw = str(value or "0").strip()
    raw = re.sub(r"[^0-9,.-]", "", raw)
    if not raw:
        return Decimal("0")
    if "," in raw:
        raw = raw.replace(".", "").replace(",", ".")
    try:
        return Decimal(raw)
    except InvalidOperation:
        return Decimal("0")


def parse_datetime(value: object) -> datetime | None:
    if isinstance(value, datetime):
        return value.replace(microsecond=0)
    raw = str(value or "").strip()
    if not raw:
        return None
    raw = raw.replace("T", " ")
    for fmt in (
        "%d/%m/%Y %H:%M:%S", "%d/%m/%Y %H:%M", "%Y-%m-%d %H:%M:%S",
        "%Y-%m-%d %H:%M", "%d-%m-%Y %H:%M:%S", "%d-%m-%Y %H:%M",
    ):
        try:
            return datetime.strptime(raw, fmt)
        except ValueError:
            pass
    return None


def parse_date(value: object):
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    raw = str(value or "").strip()
    for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y", "%Y/%m/%d"):
        try:
            return datetime.strptime(raw[:10], fmt).date()
        except ValueError:
            pass
    dt = parse_datetime(value)
    return dt.date() if dt else None


def parse_time(value: object):
    if isinstance(value, datetime):
        return value.time().replace(microsecond=0)
    if isinstance(value, dt_time):
        return value.replace(microsecond=0)
    if isinstance(value, (int, float)) and 0 <= float(value) < 1:
        total_seconds = round(float(value) * 86400) % 86400
        return dt_time(total_seconds // 3600, (total_seconds % 3600) // 60, total_seconds % 60)
    raw = str(value or "").strip()
    if not raw:
        return None
    dt = parse_datetime(raw)
    if dt:
        return dt.time()
    raw = raw.replace("h", ":").replace("H", ":")
    for fmt in ("%H:%M:%S", "%H:%M", "%H%M%S", "%H%M"):
        try:
            return datetime.strptime(raw[:8], fmt).time()
        except ValueError:
            pass
    return None


def parse_week(value: object, fallback_date: date | None = None) -> int | None:
    raw = str(value or "").strip()
    match = re.search(r"\b(\d{1,2})\b", raw)
    if match:
        number = int(match.group(1))
        return number if 1 <= number <= 53 else None
    return int(fallback_date.isocalendar().week) if fallback_date else None


def cep4_from_value(value: object) -> str | None:
    digits = re.sub(r"\D", "", str(value or ""))
    if not digits:
        return None
    return digits.zfill(8)[:4]


def faixa_horaria(value: dt_time | None) -> str | None:
    if value is None:
        return None
    minutes = value.hour * 60 + value.minute
    if minutes < 450:
        return "Antes de 07:30"
    if minutes < 600:
        return "07:30–09:59"
    if minutes < 720:
        return "10:00–11:59"
    if minutes < 840:
        return "12:00–13:59"
    if minutes < 960:
        return "14:00–15:59"
    if minutes < 1080:
        return "16:00–17:59"
    return "18:00 ou mais"


def read_rows(path: Path) -> tuple[list[str], list[dict[str, object]]]:
    if path.suffix.lower() == ".csv":
        data = path.read_bytes()
        text = data.decode("utf-8-sig", errors="replace")
        first_line = text.splitlines()[0] if text.splitlines() else ""
        # O arquivo oficial contém vírgulas em endereços e valores; por isso,
        # um CSV com cabeçalho separado por ';' deve sempre manter ';'.
        if first_line.count(";") >= 2:
            delimiter = ";"
        elif first_line.count("\t") >= 2:
            delimiter = "\t"
        else:
            delimiter = ","
        reader = csv.DictReader(io.StringIO(text), delimiter=delimiter)
        return [str(h or "").strip() for h in (reader.fieldnames or [])], list(reader)

    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    values = ws.iter_rows(values_only=True)
    try:
        headers = [str(v or "").strip() for v in next(values)]
    except StopIteration:
        wb.close()
        return [], []
    rows = [dict(zip(headers, row)) for row in values]
    wb.close()
    return headers, rows


def _backup_sqlite(label: str) -> Path | None:
    if not str(db.engine.url).startswith("sqlite"):
        return None
    source = Path(db.engine.url.database or "")
    if not source.exists():
        return None
    backup_dir = Path(current_app.root_path).parent / "backups"
    backup_dir.mkdir(parents=True, exist_ok=True)
    target = backup_dir / f"flip_{label}_{datetime.now():%Y%m%d_%H%M%S}.db"
    shutil.copy2(source, target)
    return target


def _can_delete_lote(lote: ImportacaoLote) -> bool:
    if current_user.can_view_all_bases:
        return True
    return current_user.perfil == "GERENTE_BASE" and lote.base_id == current_user.base_id


def visible_lotes_query():
    query = db.select(ImportacaoLote)
    if not current_user.can_view_all_bases:
        query = query.where(ImportacaoLote.base_id == current_user.base_id)
    return query


def _save_upload(arquivo) -> tuple[str, Path]:
    upload_dir = Path(current_app.config["UPLOAD_FOLDER"])
    upload_dir.mkdir(parents=True, exist_ok=True)
    stored_name = f"{datetime.now():%Y%m%d_%H%M%S_%f}_{secure_filename(arquivo.filename)}"
    path = upload_dir / stored_name
    arquivo.save(path)
    return stored_name, path


def _field(row: dict[str, object], mapping: dict[str, str], name: str) -> object:
    column = mapping.get(name)
    return row.get(column) if column else None


def _quality_summary(rows: list[dict[str, object]], mapping: dict[str, str]) -> dict[str, object]:
    total = len(rows)
    def count(field: str) -> int:
        col = mapping.get(field)
        return sum(1 for row in rows if col and str(row.get(col, "") or "").strip())
    scores = [
        count("tbr"), count("cliente"), count("endereco"), count("motorista"),
        count("data_dnr") or count("data_hora_entrega"), count("hora_dnr") or count("data_hora_entrega"),
    ]
    possible = total * len(scores)
    quality = round(sum(scores) * 100 / possible, 1) if possible else 0
    values = [parse_decimal(_field(row, mapping, "valor")) for row in rows]
    dates = [parse_date(_field(row, mapping, "data_dnr")) or (parse_datetime(_field(row, mapping, "data_hora_entrega")) or datetime.min).date() for row in rows]
    dates = [d for d in dates if d != datetime.min.date()]
    weeks = [parse_week(_field(row, mapping, "semana"), dates[0] if dates else None) for row in rows]
    weeks = sorted({w for w in weeks if w})
    return {
        "total": total,
        "motoristas": count("motorista"),
        "logins": count("login_utilizado"),
        "enderecos": count("endereco"),
        "datas": max(count("data_dnr"), count("data_hora_entrega")),
        "horas": max(count("hora_dnr"), count("data_hora_entrega")),
        "valor_total": sum(values, Decimal("0")),
        "qualidade": quality,
        "semanas": weeks,
    }


@bp.route("/modelo-oficial")
@login_required
def modelo_oficial():
    model_dir = Path(current_app.root_path).parent / "modelos"
    return send_from_directory(model_dir, "Modelo_Oficial_DNR_Intelligence.xlsx", as_attachment=True)


@bp.route("/")
@login_required
def index():
    lotes = db.session.scalars(visible_lotes_query().order_by(ImportacaoLote.criado_em.desc())).all()
    bases = db.session.scalars(db.select(BaseOperacional).where(BaseOperacional.ativa.is_(True)).order_by(BaseOperacional.codigo)).all()
    if not current_user.can_view_all_bases:
        bases = [current_user.base]
    return render_template("imports/index.html", lotes=lotes, bases=bases)


@bp.route("/previsualizar", methods=["POST"])
@login_required
def previsualizar():
    arquivo = request.files.get("arquivo")
    base_id = request.form.get("base_id", type=int) if (current_user.can_view_all_bases) else current_user.base_id
    if not arquivo or not arquivo.filename or not base_id:
        flash("Selecione um arquivo CSV/XLSX e uma base.", "warning")
        return redirect(url_for("imports.index"))
    ext = Path(arquivo.filename).suffix.lower()
    if ext not in {".csv", ".xlsx"}:
        flash("Formato inválido. Use CSV ou XLSX.", "warning")
        return redirect(url_for("imports.index"))

    stored_name, path = _save_upload(arquivo)
    try:
        headers, rows = read_rows(path)
        mapping = detect_columns(headers)
        if "tbr" not in mapping or "cliente" not in mapping:
            raise ValueError("A planilha precisa conter as colunas TBR e Cliente.")
        summary = _quality_summary(rows, mapping)
        preview_rows = rows[:8]
        base = db.session.get(BaseOperacional, base_id)
        return render_template(
            "imports/preview.html", original_name=arquivo.filename, stored_name=stored_name,
            base=base, headers=headers, rows=preview_rows, mapping=mapping, summary=summary,
        )
    except Exception as exc:
        path.unlink(missing_ok=True)
        flash(f"Não foi possível ler a planilha: {exc}", "warning")
        return redirect(url_for("imports.index"))


@bp.route("/confirmar", methods=["POST"])
@login_required
def confirmar():
    stored_name = secure_filename(request.form.get("stored_name", ""))
    original_name = request.form.get("original_name", "planilha")
    base_id = request.form.get("base_id", type=int)
    if not stored_name or not base_id:
        flash("Pré-visualização expirada. Selecione o arquivo novamente.", "warning")
        return redirect(url_for("imports.index"))
    if not (current_user.can_view_all_bases):
        base_id = current_user.base_id
    path = Path(current_app.config["UPLOAD_FOLDER"]) / stored_name
    if not path.exists():
        flash("O arquivo temporário não foi encontrado. Selecione-o novamente.", "warning")
        return redirect(url_for("imports.index"))
    return _process_import(path, stored_name, original_name, base_id)


# Compatibilidade com versões anteriores e automações que postavam em /novo.
@bp.route("/novo", methods=["POST"])
@login_required
def novo():
    arquivo = request.files.get("arquivo")
    base_id = request.form.get("base_id", type=int) if (current_user.can_view_all_bases) else current_user.base_id
    if not arquivo or not arquivo.filename or not base_id:
        flash("Selecione um arquivo CSV/XLSX e uma base.", "warning")
        return redirect(url_for("imports.index"))
    stored_name, path = _save_upload(arquivo)
    return _process_import(path, stored_name, arquivo.filename, base_id)



def _apply_row_to_case(caso: CasoDNR, row: dict[str, object], mapping: dict[str, str]) -> bool:
    """Atualiza um caso existente com os dados disponíveis no arquivo original."""
    changed = False
    full_dt = parse_datetime(_field(row, mapping, "data_hora_entrega"))
    data_abertura_dnr = parse_date(_field(row, mapping, "data_abertura_dnr"))
    data_dnr = parse_date(_field(row, mapping, "data_dnr")) or (full_dt.date() if full_dt else None)
    hora_dnr = parse_time(_field(row, mapping, "hora_dnr")) or (full_dt.time() if full_dt else None)
    values = {
        "cliente": str(_field(row, mapping, "cliente") or "").strip(),
        "endereco": str(_field(row, mapping, "endereco") or "").strip(),
        "cep": str(_field(row, mapping, "cep") or "").strip(),
        "cep4": cep4_from_value(_field(row, mapping, "cep")),
        "motorista": str(_field(row, mapping, "motorista") or "").strip(),
        "login_utilizado": str(_field(row, mapping, "login_utilizado") or "").strip(),
        "proprietario_login": str(_field(row, mapping, "proprietario_login") or "").strip(),
        "produto": str(_field(row, mapping, "produto") or "").strip(),
        "categoria": str(_field(row, mapping, "categoria") or "").strip(),
        "pedido": str(_field(row, mapping, "pedido") or "").strip(),
        "valor": parse_decimal(_field(row, mapping, "valor")),
        "data_abertura_dnr": data_abertura_dnr,
        "data_dnr": data_dnr,
        "hora_dnr": hora_dnr,
        "data_hora_entrega": full_dt or (datetime.combine(data_dnr, hora_dnr) if data_dnr and hora_dnr else None),
        "semana_numero": parse_week(_field(row, mapping, "semana"), data_dnr),
        "ano": data_dnr.year if data_dnr else None,
        "mes": data_dnr.month if data_dnr else None,
        "dia": data_dnr.day if data_dnr else None,
        "dia_semana": data_dnr.strftime("%A") if data_dnr else None,
        "faixa_horaria": faixa_horaria(hora_dnr),
    }
    login_raw = str(_field(row, mapping, "login_proprio") or "").strip().upper()
    if login_raw:
        values["login_proprio"] = login_raw in {"SIM", "S", "YES", "1", "TRUE"}
    for field, value in values.items():
        if value not in (None, "") and getattr(caso, field, None) != value:
            setattr(caso, field, value)
            changed = True
    return changed


def _refresh_lote(lote: ImportacaoLote) -> tuple[int, int, int]:
    path = Path(current_app.config["UPLOAD_FOLDER"]) / lote.arquivo_salvo
    if not path.exists():
        raise FileNotFoundError("O arquivo original deste lote não está mais na pasta uploads.")
    headers, rows = read_rows(path)
    mapping = detect_columns(headers)
    updated = missing = unchanged = 0
    for row in rows:
        tbr = str(_field(row, mapping, "tbr") or "").strip().upper()
        if not tbr:
            continue
        caso = db.session.scalar(db.select(CasoDNR).where(and_(CasoDNR.base_id == lote.base_id, CasoDNR.tbr == tbr)))
        if not caso:
            missing += 1
            continue
        if _apply_row_to_case(caso, row, mapping):
            updated += 1
        else:
            unchanged += 1
    summary = _quality_summary(rows, mapping)
    lote.mapeamento = json.dumps(mapping, ensure_ascii=False)
    lote.valor_total = summary["valor_total"]
    lote.qualidade_percentual = summary["qualidade"]
    lote.enderecos_preenchidos = summary["enderecos"]
    lote.motoristas_preenchidos = summary["motoristas"]
    lote.logins_preenchidos = summary["logins"]
    lote.datas_preenchidas = summary["datas"]
    lote.horas_preenchidas = summary["horas"]
    lote.semana_numero = summary["semanas"][0] if len(summary["semanas"]) == 1 else lote.semana_numero
    return updated, missing, unchanged

def _process_import(path: Path, stored_name: str, original_name: str, base_id: int):
    started = time.perf_counter()
    lote = ImportacaoLote(
        nome_arquivo=original_name, arquivo_salvo=stored_name, base_id=base_id,
        usuario_id=current_user.id, status="PROCESSANDO",
    )
    db.session.add(lote)
    db.session.commit()
    erros: list[str] = []

    try:
        headers, rows = read_rows(path)
        mapping = detect_columns(headers)
        if "tbr" not in mapping or "cliente" not in mapping:
            raise ValueError("A planilha precisa conter ao menos as colunas TBR e Cliente.")

        lote.total_linhas = len(rows)
        lote.mapeamento = json.dumps(mapping, ensure_ascii=False)
        summary = _quality_summary(rows, mapping)
        lote.valor_total = summary["valor_total"]
        lote.qualidade_percentual = summary["qualidade"]
        lote.enderecos_preenchidos = summary["enderecos"]
        lote.motoristas_preenchidos = summary["motoristas"]
        lote.logins_preenchidos = summary["logins"]
        lote.datas_preenchidas = summary["datas"]
        lote.horas_preenchidas = summary["horas"]
        lote.semana_numero = summary["semanas"][0] if len(summary["semanas"]) == 1 else None

        for idx, row in enumerate(rows, start=2):
            tbr = str(_field(row, mapping, "tbr") or "").strip().upper()
            cliente = str(_field(row, mapping, "cliente") or "").strip()
            if not tbr or not cliente:
                lote.ignorados += 1
                erros.append(f"Linha {idx}: TBR ou cliente vazio.")
                continue
            existente = db.session.scalar(db.select(CasoDNR).where(and_(CasoDNR.base_id == base_id, CasoDNR.tbr == tbr)))
            if existente:
                lote.duplicados += 1
                continue

            full_dt = parse_datetime(_field(row, mapping, "data_hora_entrega"))
            data_abertura_dnr = parse_date(_field(row, mapping, "data_abertura_dnr"))
            data_dnr = parse_date(_field(row, mapping, "data_dnr")) or (full_dt.date() if full_dt else None)
            hora_dnr = parse_time(_field(row, mapping, "hora_dnr")) or (full_dt.time() if full_dt else None)
            semana = parse_week(_field(row, mapping, "semana"), data_dnr)
            ultimo = db.session.scalar(db.select(db.func.max(CasoDNR.id))) or 0
            codigo = f"CASO-{date.today().year}-{ultimo + 1:06d}"

            login_raw = str(_field(row, mapping, "login_proprio") or "").strip().upper()
            caso = CasoDNR(
                codigo=codigo,
                tbr=tbr,
                cliente=cliente,
                endereco=str(_field(row, mapping, "endereco") or "").strip(),
                cep=str(_field(row, mapping, "cep") or "").strip(),
                cep4=cep4_from_value(_field(row, mapping, "cep")),
                motorista=str(_field(row, mapping, "motorista") or "").strip(),
                login_utilizado=str(_field(row, mapping, "login_utilizado") or "").strip(),
                login_proprio=(login_raw in {"SIM", "S", "YES", "1", "TRUE"}) if login_raw else None,
                proprietario_login=str(_field(row, mapping, "proprietario_login") or "").strip(),
                produto=str(_field(row, mapping, "produto") or "").strip(),
                categoria=str(_field(row, mapping, "categoria") or "").strip(),
                pedido=str(_field(row, mapping, "pedido") or "").strip(),
                valor=parse_decimal(_field(row, mapping, "valor")),
                data_abertura_dnr=data_abertura_dnr,
                data_dnr=data_dnr,
                hora_dnr=hora_dnr,
                data_hora_entrega=full_dt or (datetime.combine(data_dnr, hora_dnr) if data_dnr and hora_dnr else None),
                semana_numero=semana,
                ano=data_dnr.year if data_dnr else None,
                mes=data_dnr.month if data_dnr else None,
                dia=data_dnr.day if data_dnr else None,
                dia_semana=data_dnr.strftime("%A") if data_dnr else None,
                faixa_horaria=faixa_horaria(hora_dnr),
                status="PENDENTE",
                prioridade="MEDIA",
                prazo=date.today() + timedelta(days=3),
                base_id=base_id,
                importacao_id=lote.id,
            )
            db.session.add(caso)
            db.session.flush()
            lote.importados += 1

        lote.status = "CONCLUIDO"
        lote.erros = json.dumps(erros[:200], ensure_ascii=False)
        lote.tempo_processamento_ms = int((time.perf_counter() - started) * 1000)
        db.session.commit()
        flash(
            f"Importação concluída: {lote.importados} novos, {lote.duplicados} duplicados, "
            f"{lote.ignorados} ignorados. Qualidade dos dados: {lote.qualidade_percentual:.1f}%.",
            "success",
        )
    except Exception as exc:
        db.session.rollback()
        lote = db.session.get(ImportacaoLote, lote.id)
        lote.status = "ERRO"
        lote.erros = json.dumps([str(exc)], ensure_ascii=False)
        lote.tempo_processamento_ms = int((time.perf_counter() - started) * 1000)
        db.session.commit()
        flash(f"Falha na importação: {exc}", "warning")
    return redirect(url_for("imports.detalhe", lote_id=lote.id))



@bp.route("/<int:lote_id>/atualizar-dados", methods=["POST"])
@login_required
def atualizar_dados(lote_id: int):
    lote = db.session.get(ImportacaoLote, lote_id)
    if not lote:
        abort(404)
    if not (current_user.can_view_all_bases or lote.base_id == current_user.base_id):
        abort(403)
    if lote.status == "EXCLUIDO":
        flash("Lotes excluídos não podem ser atualizados.", "warning")
        return redirect(url_for("imports.index"))
    try:
        _backup_sqlite(f"antes_atualizar_lote_{lote.id}")
        updated, missing, unchanged = _refresh_lote(lote)
        db.session.commit()
        flash(f"Lote atualizado: {updated} casos corrigidos, {unchanged} já estavam atualizados e {missing} TBR não encontrados.", "success")
    except Exception as exc:
        db.session.rollback()
        flash(f"Não foi possível atualizar o lote: {exc}", "warning")
    return redirect(url_for("imports.index"))


@bp.route("/atualizar-todos", methods=["POST"])
@login_required
def atualizar_todos():
    query = visible_lotes_query().where(ImportacaoLote.status == "CONCLUIDO")
    lotes = db.session.scalars(query.order_by(ImportacaoLote.id)).all()
    total_updated = total_missing = arquivos_ausentes = 0
    _backup_sqlite("antes_atualizar_todos_lotes")
    for lote in lotes:
        try:
            updated, missing, _ = _refresh_lote(lote)
            total_updated += updated
            total_missing += missing
        except FileNotFoundError:
            arquivos_ausentes += 1
    db.session.commit()
    flash(f"Atualização concluída: {total_updated} casos corrigidos, {total_missing} TBR não encontrados e {arquivos_ausentes} arquivos originais ausentes.", "success" if total_updated else "warning")
    return redirect(url_for("imports.index"))

@bp.route("/<int:lote_id>")
@login_required
def detalhe(lote_id: int):
    lote = db.session.get(ImportacaoLote, lote_id)
    if not lote:
        return redirect(url_for("imports.index"))
    if not (current_user.can_view_all_bases) and lote.base_id != current_user.base_id:
        return redirect(url_for("imports.index"))
    casos = db.session.scalars(db.select(CasoDNR).where(CasoDNR.importacao_id == lote.id).order_by(CasoDNR.id.desc()).limit(200)).all()
    erros = json.loads(lote.erros or "[]")
    mapping = json.loads(lote.mapeamento or "{}")
    return render_template("imports/detail.html", lote=lote, casos=casos, erros=erros, mapping=mapping)


@bp.route("/<int:lote_id>/excluir", methods=["POST"])
@login_required
def excluir(lote_id: int):
    lote = db.session.get(ImportacaoLote, lote_id)
    if not lote:
        abort(404)
    if not _can_delete_lote(lote):
        abort(403)
    if lote.status == "EXCLUIDO":
        flash("Este lote já foi excluído.", "warning")
        return redirect(url_for("imports.detalhe", lote_id=lote.id))

    confirmacao = request.form.get("confirmacao", "").strip().upper()
    motivo = request.form.get("motivo", "").strip()
    if confirmacao != "EXCLUIR":
        flash("Digite EXCLUIR para confirmar a remoção do lote.", "warning")
        return redirect(url_for("imports.detalhe", lote_id=lote.id))
    if len(motivo) < 5:
        flash("Informe o motivo da exclusão com pelo menos 5 caracteres.", "warning")
        return redirect(url_for("imports.detalhe", lote_id=lote.id))

    _backup_sqlite(f"antes_excluir_lote_{lote.id}")
    casos = db.session.scalars(db.select(CasoDNR).where(CasoDNR.importacao_id == lote.id)).all()
    quantidade = len(casos)
    for caso in casos:
        db.session.delete(caso)

    lote.status = "EXCLUIDO"
    lote.excluido_em = datetime.now().astimezone()
    lote.excluido_por_id = current_user.id
    lote.motivo_exclusao = motivo
    lote.casos_excluidos = quantidade

    file_path = Path(current_app.config["UPLOAD_FOLDER"]) / lote.arquivo_salvo
    file_path.unlink(missing_ok=True)
    db.session.commit()
    flash(f"Lote excluído com segurança. {quantidade} caso(s) foram removidos.", "success")
    return redirect(url_for("imports.index"))
